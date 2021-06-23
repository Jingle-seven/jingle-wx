# coding: utf-8
# 统计2020年广东省考县以上招录的专业数量

import openpyxl
fileName = '../../resource/广东省县级以上机关2020年招录公务员职位表.xlsx'
book = openpyxl.load_workbook(fileName)
rawMajorSheet = book['原专业目录']
majorSheet = book['专业目录']
positionSheet = book['县以上机关']

codeToMajor = dict()
for rowNum in range(1, rawMajorSheet.max_row + 1):
    code = rawMajorSheet.cell(rowNum, 2).value
    major = rawMajorSheet.cell(rowNum, 3).value
    codeToMajor[code] = major
    print(code,major)
    print(codeToMajor)