# coding: utf-8
# 每月农保待遇核定名单，用来通知到龄人员来开始领取待遇，txt转excel，删掉无效人名

import openpyxl,time
import jingle.demo.work.SkData as SkData

tableHead = ['序号','村委','姓名','公民身份号码','性别','已缴费月数','当前缴费档次','是否购买职工社保']
monthStr = time.strftime("%Y年%m月", time.localtime())
xlsName = monthStr + '农保待遇名单'
dir = 'C:/Users/Administrator/Desktop/农保/农保每月待遇核定/'
rawXsl = openpyxl.load_workbook(dir + monthStr + '导出.xlsx')
sheet0 = rawXsl['sheet0']

wbook = openpyxl.Workbook()
# sheet = wbook.create_sheet('sheet1'，0) # 新建工作表（表名，位置）
# for sheet in wb:
resSheet = wbook['Sheet'] # 使用默认的Sheet
resSheet.title = xlsName # 更改表名
resSheet.append(tableHead)
resSheet2 = wbook.create_sheet('sheet1',1)
resSheet2.append(tableHead)

rowNum = 2
paidCount = 1
notPaidCount = 1
for x in sheet0.values:
    # print(x)
    if x[4] == '未领取' and x[5] == '参保缴费': #
        rowNum  = rowNum + 1
        sex = '女' if (int(x[1][-2]))/2==0 else '男'
        row = [rowNum-2,x[7].replace('村委会',''),x[3],x[1][0:-6]+'***',sex,'','','否']
        print(row)
        resSheet.append(row)
        resSheet2.append([rowNum-2,x[7].replace('村委会',''),x[3],x[1],sex,'','','否'])

SkData.setBorderWidth(resSheet, 25) # 设置边框直到25行
SkData.setBorderWidth(resSheet2, 25) # 设置边框直到25行
wbook.save(dir + xlsName + '.xlsx')