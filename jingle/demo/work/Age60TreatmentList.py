# coding: utf-8
# 每月农保待遇核定名单，用来通知到龄人员来开始领取待遇，txt转excel，删掉无效人名

import openpyxl
import jingle.demo.work.SkData as SkData

tableHead = ['序号','村委','姓名','公民身份号码','性别','已缴费月数','当前缴费档次','是否购买职工社保']
escapeNames = ['戴金兰','谭大妹','钟达海','陈如章','刘荣','李英山','戴东北','蓝善付']
hadNames = ['吴林凤','何文鲜','谢加贤','刘洪疆']
xlsName = '21年2月农保待遇核定'
dir = 'C:/Users/Administrator/Desktop/农保/农保每月待遇核定/'
txtPath = dir + xlsName + '.txt'
txtFile = open(txtPath,encoding='gbk')

wbook = openpyxl.Workbook()
# sheet = wbook.create_sheet('sheet1'，0) # 新建工作表（表名，位置）
# for sheet in wb:
print(wbook.sheetnames)
paidS = wbook['Sheet'] # 使用默认的Sheet
paidS.title = '已交清' # 更改表名
notPaidS = wbook.create_sheet('未交清')

paidS.append(tableHead)
notPaidS.append(tableHead)

rowNum = 2
paidCount = 1
notPaidCount = 1
for line in txtFile:
    x = line.split('|')
    # print(x)
    if line.startswith('#'): #第一行不解析
        continue
    if x[15] == '0':#删掉已交0期的数据
        print(x[5] + ' 去除，已交0期')
        continue
    if  x[5] in escapeNames :#删掉无效人名数据
        print(x[5] + ' 去除，无效人名')
        continue
    if rowNum == 2:#表头行不解析
        rowNum  = rowNum + 1
        continue
    row = [rowNum-2,x[4].replace('南雄市水口镇',''),x[5],x[6],x[7],x[15],x[17],'否']
    print(row)
    birthYear = int(row[3][6:10])
    monthToBePaid = (birthYear - 1950) * 12
    # print(row[3][6:10],monthToBePaid,row[5] == str(monthToBePaid))
    if row[5] == str(monthToBePaid):
        row[0] = paidCount
        paidCount = paidCount +1
        paidS.append(row)
    else:
        row[0] = notPaidCount
        notPaidCount = notPaidCount + 1
        notPaidS.append(row)
    rowNum  = rowNum + 1

SkData.setBorderWidth(paidS, 25) # 设置边框直到25行
SkData.setBorderWidth(notPaidS, 25)
wbook.save(dir + xlsName + '.xlsx')