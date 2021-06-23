# coding: utf-8
# 统计2020年广东省考县以上招录的专业数量

import openpyxl
fileName = '../../resource/广东省县级以上机关2020年招录公务员职位表.xlsx'
book = openpyxl.load_workbook(fileName)
rawMajorSheet = book['原专业目录']
majorSheet = book['专业目录']
positionSheet = book['县以上机关']

for rowNum in range(1, rawMajorSheet.max_row + 1):
    code = rawMajorSheet.cell(rowNum, 2).value
    major = rawMajorSheet.cell(rowNum, 3).value

    print(code,major)

codeToMajor = dict()
for rowNum in range(1, majorSheet.max_row + 1): # 读专业目录表，构建专业代码键值对
    code = majorSheet.cell(rowNum, 1).value
    major = majorSheet.cell(rowNum, 2).value
    majorCode = majorSheet.cell(rowNum, 3).value
    codeToMajor[majorCode] = [major, 0] # 定义键为'专业名(专业代码)'，值为专业名和累计数的列表
    # majorSheet.cell(rowNum, 3).value = majorNameCellValue + '(' + codeCellValue + ')'
codeToMajor['不限'] = ['不限',0]
for rowNum in range(1, positionSheet.max_row + 1): # 读职位表，统计其中出现的专业代码数量
    ms = positionSheet.cell(rowNum,10).value
    if ms:# 岗位要求如果不为空
        msList = ms.split(',')
        for m in msList:
            m = m.strip() # 去空格
            if m in codeToMajor:
                codeToMajor[m][1] += 1
            else:
                print(m,'不在专业目录中')

count = 1
for k,v in codeToMajor.items(): # 写下结果
    # if v[1] >0:
    #     print(k,v[1])
    majorSheet.cell(count, 5).value = k
    majorSheet.cell(count, 6).value = v[1]
    count +=1

book.save(fileName)